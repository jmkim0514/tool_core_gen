# -*- coding: utf-8 -*-
# =================================================================
# Copyright (C) 2020 AlphaHondings Co. ,All Rights Reserved
# AlphaHoldings Co. Proprietary & Confidential
# File Name   : alpgen.py 
# Description :
# Author      : Kim Jong Min ( jimmy@alpha-holdings.kr )
# =================================================================
import sys
import os
import json
import string
from pprint import pprint
#from operator import itemgetter, attrgetter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from alpgen_util import alpgen_vip, alpgen_print, alpgen_lint, alpgen_group
import copy

# alpgen
# alpgen_json
# alpgen_excel
    
#------------------------------------------------------------------------------
# alpgen_json - @mark
#------------------------------------------------------------------------------
class alpgen_json:
    def __init__ (self):
        self.JSON = {}
        self.init_json()
        self.json_file_name = ''

    def init_json(self):
        if   (self.JSON.get('top','no_key')=='no_key'):
            self.JSON['top'] = {'modname':'', 'ports':[]}
        else:
            if self.JSON['top'].get('modname','no_key')=='no_key': self.JSON['top']['modname'] = ''
            if self.JSON['top'].get('ports'  ,'no_key')=='no_key': self.JSON['top']['ports']   = []
        
        if (self.JSON.get('module'    ,'no_key')=='no_key'): self.JSON['module'] = {}
        if (self.JSON.get('instance'  ,'no_key')=='no_key'): self.JSON['instance'] = {}
        if (self.JSON.get('connection','no_key')=='no_key'): self.JSON['connection'] = []

    def write_json(self, file_name):
        self.json_file_name = file_name
        self.init_json()
        with open(self.json_file_name, 'w', encoding='utf-8') as make_file:
            json.dump(self.JSON, make_file, indent='\t')

    def read_json(self, file_name):
        self.json_file_name = file_name
        with open (self.json_file_name, 'r') as f:
            self.JSON = json.load(f)
        self.init_json()


    def remove_unused_module(self):
        used_module = []
        unused_module = []
        for inst_name in self.JSON['instance']:
            used_module.append(self.JSON['instance'][inst_name]['modname'])
        for mod_name in self.JSON['module']:
            if mod_name not in used_module:
                unused_module.append(mod_name)
        for mod_name in unused_module:
            del self.JSON['module'][mod_name]

    # Port Format Conversion - list to dictionary
    def l2d(self, lst):
        #return {'port': lst[0].strip(), 'dir':lst[1].strip(), 'msb': int(lst[2]), 'lsb': int(lst[3])}
        return {'port': lst[0].strip(), 'dir':lst[1].strip(), 'msb': lst[2], 'lsb': lst[3]}

    # Port Generator
    # input : port_name, direction, msb, lsb
    # return : {'port':port_name, 'dir': dir, 'msb': msb, 'lsb': lsb}
    def gen_port (self, port, direct, msb, lsb):
        return {'port':port.strip(), 'msb':msb, 'lsb':lsb, 'dir':direct.strip()}

    # Port Generator
    # input : [ port_name, dir, msb, lsb]
    # return : {'port':port_name, 'dir': dir, 'msb': msb, 'lsb': lsb}
    def gen_port (self, lst):
        return {'port':lst[0].strip(), 'dir':lst[1].strip(), 'msb':int(lst[2]), 'lsb':int(lst[3])}

    def add_module (self, mod_name, file_name, port_list):
        dic = {}
        dic['modname'] = mod_name
        dic['filename'] = file_name
        dic['ports'] = []
        for i in port_list:
            if   type(i) is type([]): dic['ports'].append(self.gen_port(i))
            elif type(i) is type({}): dic['ports'].append(i)
            else:
                dicert(0), '[ERROR] alpgen_json.add_module : unknown port list = %s'%(i)
        self.JSON['module'][mod_name] = dic

    def add_top (self, mod_name, port_list):
        dic = {}
        dic['modname'] = mod_name
        dic['ports'] = []
        for i in port_list:
            if   type(i) is type([]): dic['ports'].append(self.gen_port(i))
            elif type(i) is type({}): dic['ports'].append(i)
            else:
                assert(0), '[ERROR] alpgen_json.add_top : unknown port list = %s'%(i)
        self.JSON['top'] = dic

    def add_instance (self, inst_name, mod_name, uniq_name, dic_param):
        inst_name = inst_name.strip()
        self.JSON['instance'][inst_name] = {}
        self.JSON['instance'][inst_name]['modname']   = mod_name.strip()
        self.JSON['instance'][inst_name]['prefix']    = uniq_name.strip()
        self.JSON['instance'][inst_name]['parameter'] = dic_param

#    def add_instance_dic (self, inst_name, dic_inst):
#        self.JSON['instance'][inst_name] = dic_inst

#    def append_top_port(self, port, dir, msb=0, lsb=0):
#        self.JSON['top']['ports'].append({'port':port, 'msb':msb, 'lsb':lsb, 'dir':dir})

    # modify - 2022_02-23
    def append_top_port(self, port_i, dir_i='input', msb_i=0, lsb_i=0):
        flag = False
        if type(port_i) in [tuple, list]:
            for add_port in port_i:
                set_add = set(add_port.items())
                for cur_port in self.JSON['top']['ports']:
                    set_cur = set(cur_port.items())
                    if len(dict(set_cur ^ set_add))==0:
                        flag = True
                        break
        else:
            if type(port_i) in [dict]:
                add_port = {'port':port_i['port'], 'dir':port_i['dir'], 'msb':port_i['msb'], 'lsb':port_i['lsb']}
            else:
                add_port = {'port':port_i, 'msb':msb_i, 'lsb':lsb_i, 'dir':dir_i}
            set_add = set(add_port.items())
            for cur_port in self.JSON['top']['ports']:
                set_cur = set(cur_port.items())
                if len(dict(set_cur ^ set_add))==0:
                    flag = True
                    break

        if not flag:
            self.JSON['top']['ports'].append(add_port)


#            self.JSON['top']['ports'].append({'port':port, 'msb':msb, 'lsb':lsb, 'dir':dir})

    # def append_top_port(self, list_ports):
    #     for add_port in list_ports:
    #         flag = False
    #         set_add = set(add_port.items())
    #         for cur_port in self.JSON['top']['ports']:
    #             set_cur = set(cur_port.items())
    #             if len(dict(set_cur ^ set_add))==0:
    #                 flag = True
    #                 break
    #         if not flag:
    #             self.JSON['top']['ports'].append(add_port)

    def _diff(self, c0_i, c1_i):
        if len(set(c0_i.items()) ^ set(c1_i.items()))==0:
            return True
        else:
            return False

    def append_connection(self, con_i):
        flag = False
        for di in self.JSON['connection']:
            if di['type']==con_i['type']:
                if di['type']=='p2p':
                    a = 'mst'
                    b = 'slv'
                    if self._diff(di[a], con_i[a]) and self._diff(di[b], con_i[b]):
                        flag = True
                        break
                    if self._diff(di[a], con_i[b]) and self._diff(di[b], con_i[a]):
                        flag = True
                        break
                elif di['type']=='p2t':
                    a = 'mst'
                    b = 'top'
                    if self._diff(di[a], con_i[a]) and self._diff(di[b], con_i[b]):
                        flag = True
                        break
                elif di['type']=='tie':
                    if self._diff(di['mst'], con_i['mst']) and di['value']==con_i['value']:
                        flag = True
                        break
        if not flag:
            self.JSON['connection'].append(con_i)

    def add_connection (self, dic_add):
        flag = False
        for dic_con in self.JSON['connection']:
            if dic_con['type']==dic_add['type']:
                # diff dictionary
                set_con = set(dic_con.items())
                set_add = set(dic_add.items())
                if len(dict(set_con ^ set_add))==0:
                    flag = True
                    break
        if not flag:
            self.JSON['connection'].append(dic_add)

    def append_top_port_list(self, ports_i):
        """append top port from ports(list)
        Args:
            ports_i (list): [{port:, dir:, msb:, lsb}, {}]
        """
        for di in ports_i:
            self.JSON['top']['ports'].append(di)

    # def con_pin_to_pin(self, m_inst, m_port, m_msb, m_lsb, s_inst, s_port, s_msb, s_lsb):
    #     con = {}
    #     con['type'] = 'p2p'
    #     con['mst'] = {'inst':m_inst.strip(), 'port':m_port.strip(), 'msb':m_msb, 'lsb':m_lsb}
    #     con['slv'] = {'inst':s_inst.strip(), 'port':s_port.strip(), 'msb':s_msb, 'lsb':s_lsb}
    #     self.JSON['connection'].append(con)

    #@modify - from tools_bus_gen
    def con_pin_to_pin(self, minst_i, mport_i, mmsb_i=0, mlsb_i=0, sinst_i='', sport_i='', smsb_i=0, slsb_i=0):
        if type(minst_i) in [tuple, list]:
            mst_inst, mst_port, mst_msb, mst_lsb = minst_i
            slv_inst, slv_port, slv_msb, slv_lsb = mport_i
        else:
            if minst_i is None: mst_inst = minst_i
            else              : mst_inst = minst_i.strip()
            if sinst_i is None: slv_inst = sinst_i
            else              : slv_inst = sinst_i.strip()
            mst_port=mport_i.strip(); mst_msb=mmsb_i; mst_lsb=mlsb_i
            slv_port=sport_i.strip(); slv_msb=smsb_i; slv_lsb=slsb_i
        con = {}
        # print('--')
        # print (type(minst_i))
        # print(mst_inst)
        if mst_inst in [None, '$top', '']: #  or mst_inst.isspace():
            con['type'] = 'p2t'
            con['mst'] = {'inst':slv_inst, 'port':slv_port, 'msb':slv_msb, 'lsb':slv_lsb}
            con['top'] = {                 'port':mst_port, 'msb':mst_msb, 'lsb':mst_lsb}
        elif slv_inst in [None, '$top', '']: # or slv_inst.isspace():
            con['type'] = 'p2t'
            con['mst'] = {'inst':mst_inst, 'port':mst_port, 'msb':mst_msb, 'lsb':mst_lsb}
            con['top'] = {                 'port':slv_port, 'msb':slv_msb, 'lsb':slv_lsb}
        else:
            con['type'] = 'p2p'
            con['mst'] = {'inst':mst_inst, 'port':mst_port, 'msb':mst_msb, 'lsb':mst_lsb}
            con['slv'] = {'inst':slv_inst, 'port':slv_port, 'msb':slv_msb, 'lsb':slv_lsb}
        self.JSON['connection'].append(con)


    def con_pin_to_top(self, m_inst, m_port, m_msb, m_lsb, s_port, s_msb, s_lsb):
        con = {}
        con['type'] = 'p2t'
        con['mst'] = {'inst':m_inst.strip(), 'port':m_port.strip(), 'msb':m_msb, 'lsb':m_lsb}
        con['top'] = {                       'port':s_port.strip(), 'msb':s_msb, 'lsb':s_lsb}
        self.JSON['connection'].append(con)

    def con_pin_value(self, m_inst, m_port, m_msb, m_lsb, value):
        con = {}
        con['type'] = 'tie'
        con['mst'] = {'inst':m_inst.strip(), 'port':m_port.strip(), 'msb':m_msb, 'lsb':m_lsb}
        con['value'] = int(value)
        self.JSON['connection'].append(con)

    def con_bus_group (self, m_inst, m_sym, s_inst, s_sym):
        tmp = {}
        if m_sym.split('$')[0].lower()!=s_sym.split('$')[0].lower():
            assert(0), '[ERROR] mst_symbol=%s, slv_symbol=%s <= different bus\n'%(m_sym, s_sym)
        
        tmp['type'] = 'bus_group'
        tmp['mst']  = {'inst': m_inst, 'symbol': m_sym}
        tmp['slv']  = {'inst': s_inst, 'symbol': s_sym}
        self.JSON['connection'].append(tmp)

    def con_pattern (self, m_inst, m_sym, s_inst, s_sym):
        tmp = {}
        tmp['type'] = 'pattern'
        tmp['mst']  = {'inst': m_inst, 'symbol': m_sym}
        tmp['slv']  = {'inst': s_inst, 'symbol': s_sym}
        self.JSON['connection'].append(tmp)

    # def con_alp_vip (self, bus_inst, bus_symbol, vip_inst, vip_clock, vip_reset):
    #     tmp = {}
    #     tmp['type'] = 'alp_vip'
    #     tmp['bus']  = {'inst': bus_inst, 'symbol': bus_symbol}
    #     tmp['vip']  = {'inst': vip_inst, 'clock': vip_clock, 'reset': vip_reset}
    #     self.JSON['connection'].append(tmp)


    def con_alp_vip (self, bus_inst, bus_sym, vip_inst, clk_inst, clk_port, rst_inst, rst_port, pre_i=None):
        if clk_port in [None, '']    : vip_clock = None
        else:
            if clk_inst in [None, '']: vip_clock = clk_port
            else                     : vip_clock = clk_inst+'.'+clk_port
        if rst_port in [None, '']    : vip_reset = None
        else:
            if rst_inst in [None, '']: vip_reset = rst_port
            else                     : vip_reset = rst_inst+'.'+rst_port
        if pre_i==None or pre_i.strip()=='': prefix = None
        else                               : prefix = pre_i
        tmp = {}
        tmp['type'] = 'alp_vip'
        tmp['bus']  = {'inst': bus_inst, 'symbol': bus_sym}
        tmp['vip']  = {'inst': vip_inst, 'clock': vip_clock, 'reset': vip_reset, 'prefix':prefix}
        self.JSON['connection'].append(tmp)

    def print(self):
        pprint (self.JSON)

#------------------------------------------------------------------------------
# alpgen_excel - @mark
#------------------------------------------------------------------------------
class alpgen_excel(alpgen_group):

    def __init__ (self, file_name):
        self.excel_file_path = file_name

    def __check_empty(self, istr):
        if   istr==None:
            return None
        else:
            if type(istr)==str:
                istr = istr.strip()
                if len(istr)==0: return None
                else           : return istr
            elif type(istr)==int:
                return str(istr)
            else:
                assert(0), '[ERROR] Unknown type'

    def str_to_dic(self, i_str):
        dic = {}
        i_str = i_str.replace('{','').replace('}','')
        if len(i_str.strip())==0:
            return dic
        for i in i_str.split(','):
            key, value = i.split(':')
            dic[key.strip()] = int(value.strip())
        return dic

    def dic_to_str(self, i_dic):
        _str = ''
        for key, value in i_dic.items():
            _str = _str+key+':'+str(value)+','
        _str = '{'+_str[:-1]+'}'
        return _str

    def format_normal(self, ws_i, loc_i):
        #y_color = PatternFill(start_color='DCE2F0', end_color='DCE2F0', fill_type='solid')
        y_color = PatternFill(start_color='DCE2F0', end_color='DCE2F0', fill_type='solid')
        b_format = Side(border_style="thin", color="00C0C0C0")
        ws_i[loc_i].fill = y_color
        ws_i[loc_i].border = Border(top=b_format, left=b_format, right=b_format, bottom=b_format)

    def format_title(self, ws_i, loc_i):
        #y_color = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        y_color = PatternFill(start_color='50586C', end_color='50586C', fill_type='solid')
        b_format = Side(border_style="thin", color="00C0C0C0")
        ws_i[loc_i].fill = y_color
        ws_i[loc_i].border = Border(top=b_format, left=b_format, right=b_format, bottom=b_format)
        ws_i[loc_i].font = Font(bold=True, color='DCE2F0')
        ws_i[loc_i].alignment = Alignment(horizontal='center', vertical='center')

#B8CCE4

    # Make dic_excel to write excel file
    #  - [ {'modname': module_name,
    #         'ports': [{'dir': input/bus_symbol,
    #                    'port': port/bus_symbol,
    #                    'msb' : 0,
    #                    'lsb': 0}], }
    #def write_excel_format_0 (self, file_path, dic_json, dic_dport):
    def write_excel_format_0 (self, dic_json, top_info_i, dic_dport):
        try:
            wb = load_workbook(self.excel_file_path)
        except:
            wb = Workbook()
            
        for sheet_name in wb.get_sheet_names():
            wb.remove(wb[sheet_name])
        
        list_col = list(string.ascii_uppercase)

        # $top WorkBook
        ws = wb.create_sheet('$top')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 70
        num = 1
        ws['A'+str(num)] = "HIERARCHY"
        font_0 = ws['A'+str(num)]; font_0.font = Font(bold=True)
        num = num + 1
        if len(top_info_i)==0:
            ws['A'+str(num)] = '$'+dic_json['top']['modname']
            num = num + 1
            for inst_name in dic_json['instance']:
                mod_name = dic_json['instance'][inst_name]['modname']
                loc = 'B'+str(num); ws[loc] = inst_name#; self.format_normal(ws, loc)
                num = num + 1
        else:
            flag = False
            for lines in top_info_i:
                if lines[0]=='HIERARCHY':
                    flag = True
                    continue
                elif lines[0]=='INSTANCE':
                    break
                if flag:
                    for col, col_data in enumerate(lines):
                        # if col_data is None:
                        #     continue
                        loc = list_col[col]+str(num)
                        ws[loc] = col_data
                        #self.format_normal(ws, loc)
                    num = num + 1
            num = num - 1

        #         print(line)
        # exit()
        # $top sheet
        num = num + 1
        ws['A'+str(num)] = "INSTANCE"        ; cb = ws['A'+str(num)]; cb.font = Font(bold=True)
        num = num + 1
        loc='A'+str(num); ws[loc] = "Instance name"   ; self.format_title(ws, loc)#cb = ws['A'+str(num)]; cb.font = Font(bold=True)
        loc='B'+str(num); ws[loc] = "Module name"     ; self.format_title(ws, loc)#cc = ws['B'+str(num)]; cc.font = Font(bold=True)
        loc='C'+str(num); ws[loc] = "File path"       ; self.format_title(ws, loc)#cd = ws['C'+str(num)]; cd.font = Font(bold=True)
        loc='D'+str(num); ws[loc] = "Parameter (dict)"; self.format_title(ws, loc)#ce = ws['D'+str(num)]; ce.font = Font(bold=True)
        num = num + 1
        for inst_name in dic_json['instance']:
            mod_name = dic_json['instance'][inst_name]['modname']
            file_path = dic_json['module'][mod_name]['filename']
            parameter = self.dic_to_str(dic_json['instance'][inst_name]['parameter'])
            loc='A'+str(num); ws[loc] = inst_name; self.format_normal(ws, loc) 
            loc='B'+str(num); ws[loc] = mod_name ; self.format_normal(ws, loc)
            loc='C'+str(num); ws[loc] = file_path; self.format_normal(ws, loc)
            loc='D'+str(num); ws[loc] = parameter; self.format_normal(ws, loc)
            num = num + 1

        # INSTANCE  WorkBook
        for inst_name in dic_json['instance']:
            ws = wb.create_sheet(title = inst_name)
            num = 1
            # loc='A'+str(num); ws[loc] = 'Direction'       ; cb = ws[loc]; cb.font = Font(bold=True)
            # loc='B'+str(num); ws[loc] = 'Port name'       ; cc = ws[loc]; cc.font = Font(bold=True)
            # loc='C'+str(num); ws[loc] = 'MSB'             ; cd = ws[loc]; cd.font = Font(bold=True)
            # loc='D'+str(num); ws[loc] = 'LSB'             ; ce = ws[loc]; ce.font = Font(bold=True)
            # loc='E'+str(num); ws[loc] = 'inst_name'       ; cf = ws[loc]; cf.font = Font(bold=True)
            # loc='F'+str(num); ws[loc] = 'port_name'       ; cg = ws[loc]; cg.font = Font(bold=True)
            # loc='G'+str(num); ws[loc] = 'vip - clock_inst'; cf = ws[loc]; cf.font = Font(bold=True)
            # loc='H'+str(num); ws[loc] = 'vip - clock_port'; cg = ws[loc]; cg.font = Font(bold=True)
            # loc='I'+str(num); ws[loc] = 'vip - reset_inst'; cf = ws[loc]; cf.font = Font(bold=True)
            # loc='J'+str(num); ws[loc] = 'vip - reset_port'; cg = ws[loc]; cg.font = Font(bold=True)
            # loc='K'+str(num); ws[loc] = 'vip - prefix'    ; cg = ws[loc]; cg.font = Font(bold=True)

            loc='A'+str(num); ws[loc] = 'Direction'       ; self.format_title(ws, loc)
            loc='B'+str(num); ws[loc] = 'Port name'       ; self.format_title(ws, loc)
            loc='C'+str(num); ws[loc] = 'MSB'             ; self.format_title(ws, loc)
            loc='D'+str(num); ws[loc] = 'LSB'             ; self.format_title(ws, loc)
            loc='E'+str(num); ws[loc] = 'inst_name'       ; self.format_title(ws, loc)
            loc='F'+str(num); ws[loc] = 'port_name'       ; self.format_title(ws, loc)
            loc='G'+str(num); ws[loc] = 'vip - clock_inst'; self.format_title(ws, loc)
            loc='H'+str(num); ws[loc] = 'vip - clock_port'; self.format_title(ws, loc)
            loc='I'+str(num); ws[loc] = 'vip - reset_inst'; self.format_title(ws, loc)
            loc='J'+str(num); ws[loc] = 'vip - reset_port'; self.format_title(ws, loc)
            loc='K'+str(num); ws[loc] = 'vip - prefix'    ; self.format_title(ws, loc)

            num = 2
            mod_name = dic_json['instance'][inst_name]['modname']
            # Normal Port
            for dic_port in dic_json['module'][mod_name]['ports']:
                if dic_dport[inst_name][dic_port['port']]['bus_group']:
                    continue
                port_name = dic_port['port']
                # try   : con_inst = dic_dport[inst_name][port_name]['con_inst']
                # except: con_inst = None
                # try   : con_port = dic_dport[inst_name][port_name]['con_port']
                # except: con_port = None
                loc='A'+str(num); ws[loc] = dic_port['dir']; self.format_normal(ws, loc)
                loc='B'+str(num); ws[loc] = port_name      ; self.format_normal(ws, loc)
                loc='C'+str(num); ws[loc] = dic_port['msb']; self.format_normal(ws, loc)
                loc='D'+str(num); ws[loc] = dic_port['lsb']; self.format_normal(ws, loc)
                # if con_inst!=None: ws['E'+str(num)] = con_inst
                # if con_port!=None: ws['F'+str(num)] = con_port

                if dic_dport[inst_name][port_name].get('con_inst'):
                    ws['E'+str(num)] = dic_dport[inst_name][port_name]['con_inst']
                if dic_dport[inst_name][port_name].get('con_port'):
                    ws['F'+str(num)] = dic_dport[inst_name][port_name]['con_port']

                # loc = 'E'+str(num); self.format_normal(ws, loc)
                # loc = 'F'+str(num); self.format_normal(ws, loc)
                # loc = 'G'+str(num); self.format_normal(ws, loc)
                # loc = 'H'+str(num); self.format_normal(ws, loc)
                # loc = 'I'+str(num); self.format_normal(ws, loc)
                # loc = 'J'+str(num); self.format_normal(ws, loc)
                # loc = 'K'+str(num); self.format_normal(ws, loc)

                num = num+1
            # Bus Group
            for symbol, dic_symbol in dic_json['module'][mod_name]['symbol_info'].items():
                port_name = self.pattern_to_symbol(dic_symbol)
                loc='A'+str(num) ; ws[loc] = 'bus$'+dic_symbol['type']            ; self.format_normal(ws, loc)
                loc='B'+str(num) ; ws[loc] = port_name                            ; self.format_normal(ws, loc)
                loc='C'+str(num) ; ws[loc] = dic_symbol['parameter']['DATA_WIDTH']; self.format_normal(ws, loc)  #msb
                loc='D'+str(num) ; ws[loc] = 0                                    ; self.format_normal(ws, loc)  #lsb
                loc='E'+str(num) #; self.format_normal(ws, loc)
                loc='F'+str(num) #; self.format_normal(ws, loc)
                loc='G'+str(num) #; self.format_normal(ws, loc)
                loc='H'+str(num) #; self.format_normal(ws, loc)
                loc='I'+str(num) #; self.format_normal(ws, loc)
                loc='J'+str(num) #; self.format_normal(ws, loc)
                loc='K'+str(num) #; self.format_normal(ws, loc)
                # if dic_dport[inst_name][port_name].get('con_inst'):
                #     ws['E'+str(num)] = dic_dport[inst_name][port_name]['con_inst']
                # if dic_dport[inst_name][port_name].get('con_port'):
                #     ws['F'+str(num)] = dic_dport[inst_name][port_name]['con_port']
                # if dic_dport[inst_name][port_name].get('vip_clock_inst'):
                #     ws['G'+str(num)] = dic_dport[inst_name][port_name]['vip_clock_inst']
                # if dic_dport[inst_name][port_name].get('vip_clock_port'):
                #     ws['H'+str(num)] = dic_dport[inst_name][port_name]['vip_clock_port']
                # if dic_dport[inst_name][port_name].get('vip_reset_inst'):
                #     ws['I'+str(num)] = dic_dport[inst_name][port_name]['vip_reset_inst']
                # if dic_dport[inst_name][port_name].get('vip_reset_port'):
                #     ws['J'+str(num)] = dic_dport[inst_name][port_name]['vip_reset_port']

                try   : ws['E'+str(num)] = dic_dport[inst_name][port_name]['con_inst']
                except: pass
                try   : ws['F'+str(num)] = dic_dport[inst_name][port_name]['con_port']
                except: pass
                try   : ws['G'+str(num)] = dic_dport[inst_name][port_name]['vip_clock_inst']
                except: pass
                try   : ws['H'+str(num)] = dic_dport[inst_name][port_name]['vip_clock_port']
                except: pass
                try   : ws['I'+str(num)] = dic_dport[inst_name][port_name]['vip_reset_inst']
                except: pass
                try   : ws['J'+str(num)] = dic_dport[inst_name][port_name]['vip_reset_port']
                except: pass
                try   : ws['K'+str(num)] = dic_dport[inst_name][port_name]['vip_prefix']
                except: pass
                num = num+1

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 5
            ws.column_dimensions['D'].width = 5
            ws.column_dimensions['E'].width = 25  # inst_name
            ws.column_dimensions['F'].width = 25  # port_name
            ws.column_dimensions['G'].width = 18  # inst_name (vip_clk)
            ws.column_dimensions['H'].width = 18  # port_name (vip_clk)
            ws.column_dimensions['I'].width = 18  # inst_name (vip_rst)
            ws.column_dimensions['J'].width = 18  # port_name (vip_rst)
            ws.column_dimensions['K'].width = 12  # vip_prefix

        wb.save(filename=self.excel_file_path)


    # return : {sheet_name : [ {'dir': input/bus_group, 'port':, 'msb':, 'lsb':, 'con_inst':, 'con_port':}, . . ]}
    def read_excel_format_0 (self):
        wb = load_workbook(self.excel_file_path)
        list_ws = wb.get_sheet_names()
        result = {}
        for sheet_name in list_ws:
            ws = wb.get_sheet_by_name(sheet_name)
            col = ws.min_column

            if sheet_name=='$top':
                top_list = []
                for x in range (ws.min_row, ws.max_row+1):
                    tmp = []
                    for y in range (ws.min_column, ws.max_column+1):
                        tmp.append(ws.cell(row=x, column=y).value)
                    top_list.append(tmp)
                result[sheet_name] = top_list
            else:
                list_ws = []
                flag_1st = True
                for row in range (ws.min_row, ws.max_row+1):
                    if flag_1st:
                        flag_1st = False
                        continue
                    dic_port = {}
                    dic_port['dir']       = self.__check_empty(ws.cell(row, col  ).value)
                    dic_port['port']      = self.__check_empty(ws.cell(row, col+1).value)
                    dic_port['msb']       =                    ws.cell(row, col+2).value
                    dic_port['lsb']       =                    ws.cell(row, col+3).value
                    dic_port['con_inst']  = self.__check_empty(ws.cell(row, col+4).value)
                    dic_port['con_port']  = self.__check_empty(ws.cell(row, col+5).value)
                    dic_port['vip_clock_inst'] = self.__check_empty(ws.cell(row, col+6).value)
                    dic_port['vip_clock_port'] = self.__check_empty(ws.cell(row, col+7).value)
                    dic_port['vip_reset_inst'] = self.__check_empty(ws.cell(row, col+8).value)
                    dic_port['vip_reset_port'] = self.__check_empty(ws.cell(row, col+9).value)
                    dic_port['vip_prefix'] = self.__check_empty(ws.cell(row, col+10).value)

                    # dic_port['vip_clock'] = self.__check_empty(ws.cell(row, col+6).value)
                    # dic_port['vip_reset'] = self.__check_empty(ws.cell(row, col+7).value)


                    list_ws.append(dic_port)
                result[sheet_name] = list_ws
        return result
    
#------------------------------------------------------------------------------
# alpgen - @mark
#------------------------------------------------------------------------------
#class alpgen(alpgen_print, alpgen_lint, alpgen_group):
class alpgen(alpgen_print, alpgen_lint, alpgen_group, alpgen_json):
    #JSON = {}
    #DTOP = {}
    #TOP_WIRE_MST = []
    #TOP_WIRE_SLV = []
    # TOP_ASSIGN
    # assign = var ==> [ ['assign_name', assign_msb, assign_lsb, var_name, var_lsb, var_msb], , , ]
    # assign = tie ==> [ ['assign_name', assign_msb, assign_lsb, value], , , ]
    #TOP_ASSIGN = []
    #open_i = 0
    #open_o = 0
    WIRE_NAME_OPEN_I = 'NC_in'
    WIRE_NAME_OPEN_O = 'NC_out'
    WIRE_NAME_OPEN_B = 'NC_inout'
    FLAG_DISP_MON  = False
    FLAG_DISP_FILE = True

    def __init__ (self, json_file_name):
        self.initial()
        with open (json_file_name, 'r') as f:
            self.JSON = json.load(f)
        log_file_name = json_file_name+'.log'
        self.DEBUG = 3  #0 : off *E, 1 : off *W, 2 : off INFO, 3 : all display

        #self.log_file = open(log_file_name, 'w')
        print ('[LOG]--------------')
        print ('[LOG] json load')
        print ('[LOG]   - json file = ', json_file_name)
        #print ('JSON File Name = ', json_file_name)
        #print ('Log File Name = ', log_file_name)

    def initial (self):
        self.JSON = {}
        self.DTOP = {}
        self.TOP_WIRE_MST = []
        self.TOP_WIRE_SLV = []
        self.TOP_ASSIGN = []
        self.open_i = 0
        self.open_o = 0
        self.open_b = 0

     # @mark
    def platform_designer (self):
        self.__make_json_format()
        self.connection_alp_vip()
        self.lint_json()

        self.dtop_init()
        self.connection_bus_group()
        self.connection_pattern()
        self.dtop_set_role_con()
        self.dtop_set_wire()
        return self.DTOP

    def write_rtl(self, rtl_file_name):
        print ('[LOG] alpgen write_rtl start')
        print ('[LOG]    - write rtl file = %s'%(rtl_file_name))
        #self.lint_json_rtl(self.JSON)
        self.__make_json_format()
        self.connection_alp_vip()
        self.lint_json()
        self.dtop_init()
        self.connection_bus_group()
        self.connection_pattern()
        self.dtop_set_role_con()
        self.dtop_set_wire()
        self.print_rtl(rtl_file_name)
        #self.log_file.close()
        print ('[LOG] alpgen write_rtl finish')

    def write_excel(self, excel_file_name):
        self.lint_json_excel_group(self.JSON)
        self.dtop_init()
        self.__make_json_format()
        #self.set_bus_group()

        print ('[LOG] Write File - format 0')
        print ('[LOG]   - Excel file name : '+excel_file_name)
        print ('[LOG]   - JSON file name  : '+excel_file_name+'.json')
        #self.write_excel_format_0(excel_file_name, self.JSON, self.DTOP)
        doc = alpgen_excel(excel_file_name)
        doc.write_excel_format_0(self.JSON, [], self.DTOP)
        with open(excel_file_name+'.json', 'w', encoding='utf-8') as f:
            json.dump(self.JSON, f, indent='\t')

    def write_excel_group(self, excel_file_name):
        print ('[LOG] Write Excel - group view')
        print ('[LOG]   - excel file name : '+excel_file_name)
        print ('[LOG]   - json file name  : '+excel_file_name+'.json')
        self.lint_json_excel_group(self.JSON)
        self.__make_json_format()
        self.dtop_init()
        self.set_bus_group()
        # Update the information from existing excel doc
        top_info = []
        if os.path.isfile(excel_file_name):
            pre_read = alpgen_excel(excel_file_name)
            for inst_name, list_port in pre_read.read_excel_format_0().items():
                try   : self.DTOP[inst_name]
                except: continue
                if inst_name=='$top':
                    top_info = list_port
                    continue
                for dic_port in list_port:
                    port_name = dic_port['port']
                    try   :self.DTOP[inst_name][port_name]
                    # for bus_group port (ex: apb$$_aud_m9)
                    except:self.DTOP[inst_name][port_name] = {}
                    self.DTOP[inst_name][port_name]['con_inst']  = dic_port['con_inst']
                    self.DTOP[inst_name][port_name]['con_port']  = dic_port['con_port']
                    self.DTOP[inst_name][port_name]['vip_clock_inst'] = dic_port['vip_clock_inst']
                    self.DTOP[inst_name][port_name]['vip_clock_port'] = dic_port['vip_clock_port']
                    self.DTOP[inst_name][port_name]['vip_reset_inst'] = dic_port['vip_reset_inst']
                    self.DTOP[inst_name][port_name]['vip_reset_port'] = dic_port['vip_reset_port']
                    self.DTOP[inst_name][port_name]['vip_prefix'] = dic_port['vip_prefix']
        doc = alpgen_excel(excel_file_name)
        doc.write_excel_format_0(self.JSON, top_info, self.DTOP)
        with open(excel_file_name+'.json', 'w', encoding='utf-8') as f:
            json.dump(self.JSON, f, indent='\t')

    # only - write_excel_group function (make gruop)
    def set_bus_group (self):
        for inst_name in self.DTOP:
            for port_name in self.DTOP[inst_name]:
                if self.DTOP[inst_name][port_name]['bus_group']:
                    continue
                dic_pattern = {}
                dic_pattern = self.get_bus_pattern(port_name, self.DTOP[inst_name][port_name]['dir'])

                if   self.DTOP[inst_name][port_name]['dir']=='output': bus_type = 'mst'
                elif self.DTOP[inst_name][port_name]['dir']=='input' : bus_type = 'slv'
                else: exit()

                if len(dic_pattern)>0:
                    for dic_bus_signal in self.get_bus_signal_list(dic_pattern, bus_type):
                        for port in dic_bus_signal['port']:
                            if (self.DTOP[inst_name].get(port, 'not_found')=='not_found'):
                                continue
                            self.DTOP[inst_name][port]['bus_group'] = True
                            if len(dic_bus_signal['parameter'])>0:
                                parameter_name  = dic_bus_signal['parameter']
                                parameter_value = self.DTOP[inst_name][port]['msb'] + \
                                                  self.DTOP[inst_name][port]['lsb'] + 1
                                dic_pattern['parameter'][parameter_name] = parameter_value
                    mod_name = self.JSON['instance'][inst_name]['modname']
                    symbol_name = self.pattern_to_symbol(dic_pattern)
                    self.JSON['module'][mod_name]['symbol_info'][symbol_name] = dic_pattern

    def connection_alp_vip(self):
        num = 0
        for con in self.JSON['connection']:
            if con['type']=='alp_vip':
                bus_inst   = con['bus']['inst']
                bus_symbol = con['bus']['symbol']
                vip_inst   = con['vip']['inst']
                vip_clock  = con['vip']['clock']
                vip_reset  = con['vip']['reset']
                vip_prefix  = con['vip']['prefix']
                bus_module = self.JSON['instance'][bus_inst]['modname']
                bus_pattern = self.JSON['module'][bus_module]['symbol_info'][bus_symbol]
                bus_type = self.JSON['module'][bus_module]['symbol_info'][bus_symbol]['type']  # mst/slv
                alp_vip = alpgen_vip(bus_pattern)
                vip_module = alp_vip.get_vip_module()  #dic
                vip_symbol = alp_vip.get_vip_symbol()
                vip_modname = vip_module['modname']
                # register module
                self.JSON['module'][vip_modname] = vip_module #dic
                # register instance
                tmp = {}
                tmp['modname'] = vip_modname
                #tmp['prefix'] = 'vip'+str(num)+'_'

                tmp['prefix'] = vip_inst if vip_prefix==None else vip_prefix
                # print ('-----')
                # print(vip_inst)
                # print(vip_prefix)

                tmp['parameter'] = bus_pattern['parameter']
                self.JSON['instance'][vip_inst] = tmp
                # register connection
                #tmp = {}
                tmp['type'] = 'bus_group'
                if bus_type=='mst':
                    tmp['mst']  = {'inst': bus_inst, 'symbol': bus_symbol}
                    tmp['slv']  = {'inst': vip_inst, 'symbol': vip_symbol}
                else:
                    tmp['mst']  = {'inst': vip_inst, 'symbol': vip_symbol}
                    tmp['slv']  = {'inst': bus_inst, 'symbol': bus_symbol}
                if vip_clock!=None:
                    vip_clock_port = alp_vip.get_vip_clock()
                    if vip_clock.find('.')>-1:
                        clock_inst, clock_port = vip_clock.split('.')
                        self.con_pin_to_pin(clock_inst, clock_port, 0, 0, vip_inst, vip_clock_port, 0, 0)
                    else:
                        self.con_pin_to_top(vip_inst, vip_clock_port, 0, 0, vip_clock, 0, 0)
                        self.append_top_port(vip_clock_port, 'input', 0, 0)
                if vip_reset!=None:
                    vip_reset_port = alp_vip.get_vip_reset()
                    if vip_reset.find('.')>-1:
                        reset_inst, reset_port = vip_reset.split('.')
                        self.con_pin_to_pin(reset_inst, reset_port, 0, 0, vip_inst, vip_reset_port, 0, 0)
                    else:
                        self.con_pin_to_top(vip_inst, vip_reset_port, 0, 0, vip_reset, 0, 0)
                        self.append_top_port(vip_reset_port, 'input', 0, 0)
                self.JSON['connection'].append(tmp)
                num = num + 1

    def dtop_init (self):
        # copy modulde info to instance info
        # make $top.port dictionary
        if (self.JSON.get('top','no_key')!='no_key'):
            if (self.JSON['top'].get('ports','no_key')!='no_key'):
                remove_list = []
                tmp_port = {}
                for num, dic_port in enumerate(self.JSON['top']['ports']):
                    if (tmp_port.get(dic_port['port'],'no_key')=='no_key'):
                        tmp_port[dic_port['port']] = self.dtop_port_format(dic_port)
                    else:
                        print ('[WARNING] More then one port ==> port name : '+dic_port['port'])
                self.DTOP['$top'] = tmp_port
        # make instance_name.port dictionary
        for inst_name in self.JSON['instance']:
            mod_name   = self.JSON['instance'][inst_name]['modname']
            #try:
            #    # ??????????? why - shallow copy ??????
            #    dic_parameter = copy.copy(self.JSON['instance'][inst_name]['parameter'])
            #except KeyError:
            #    dic_parameter = {}
            tmp_port = {}
            for dic_port in self.JSON['module'][mod_name]['ports']:
                #tmp_port[dic_port['port']] = self.dtop_port_format1(dic_port, dic_parameter, inst_name)
                tmp_port[dic_port['port']] = self.dtop_port_format1(dic_port, inst_name)
            self.DTOP[inst_name] = tmp_port

    def connection_bus_group (self):
        for con in self.JSON['connection']:
            if con['type']!='bus_group': continue
            # Bus Operation
            mst_inst = con['mst']['inst']
            mst_sym  = con['mst']['symbol']
            slv_inst = con['slv']['inst']
            slv_sym  = con['slv']['symbol']
            l_mst_bus_signal = self.get_bus_signal_list( self.symbol_to_pattern(mst_sym), 'mst' )
            l_slv_bus_signal = self.get_bus_signal_list( self.symbol_to_pattern(slv_sym), 'slv' )
            hready_loopback = ''   # ahb hready connection
            dic_hready = {'mst':{'inst':'', 'input':'', 'output':''}, 'slv':{'inst':'', 'input':'', 'output':''}}

            for num in range(len(l_mst_bus_signal)):
                mst_default = l_mst_bus_signal[num]['default']
                mst_level   = l_mst_bus_signal[num]['level']
                # Search Master Port
                mst_flag = False
                mst_dir     = l_mst_bus_signal[num]['dir']
                for mst_port in l_mst_bus_signal[num]['port']:
                    # port_name check
                    # try   : self.DTOP[mst_inst][mst_port]
                    # except: continue

                    if self.DTOP[mst_inst].get(mst_port)==None: #2022-03-13
                        continue

                    # port_direction check
                    if mst_dir!=self.DTOP[mst_inst][mst_port]['dir']:
                        continue
                    # AHB hready 
                    if mst_level==10:
                        dic_hready['mst']['inst'] = mst_inst
                        if   mst_dir=='output': dic_hready['mst']['output'] = mst_port
                        elif mst_dir=='input' : dic_hready['mst']['input']  = mst_port
                    else:
                        mst_flag = True
                        mst_msb = self.DTOP[mst_inst][mst_port]['msb']
                        mst_lsb = self.DTOP[mst_inst][mst_port]['lsb']

                    break   #2022-03-13
                    #-== if mst_dir==self.DTOP[mst_inst][mst_port]['dir']:
                    #-==     mst_flag = True
                    #-==     mst_msb = self.DTOP[mst_inst][mst_port]['msb']
                    #-==     mst_lsb = self.DTOP[mst_inst][mst_port]['lsb']
                    #-==     if mst_level==10:
                    #-==         dic_hready['mst']['inst'] = mst_inst
                    #-==         if   mst_dir=='output': dic_hready['mst']['output'] = mst_port
                    #-==         elif mst_dir=='input' : dic_hready['mst']['input']  = mst_port
                    #-==     #-== if mst_level==10:
                    #-==     #-==     hready_loopback = mst_port
                    #-==     #-== break

                # Search Slave Port
                slv_flag = False
                slv_dir     = l_slv_bus_signal[num]['dir']
                for slv_port in l_slv_bus_signal[num]['port']:
                    # port name check
                    try   : self.DTOP[slv_inst][slv_port]
                    except: continue
                    # port direction check
                    if slv_dir!=self.DTOP[slv_inst][slv_port]['dir']:
                        continue
                    #slv_dir = self.DTOP[slv_inst][slv_port]['dir']
                    slv_msb = self.DTOP[slv_inst][slv_port]['msb']
                    slv_lsb = self.DTOP[slv_inst][slv_port]['lsb']
                    if mst_level==10:
                        dic_hready['slv']['inst'] = slv_inst
                        if   slv_dir=='output': dic_hready['slv']['output'] = slv_port
                        elif slv_dir=='input' : dic_hready['slv']['input']  = slv_port
                    if mst_flag:
                        if mst_dir!=slv_dir: slv_flag = True; break    
                    else:
                        if mst_dir!=slv_dir: slv_flag = True; break
                    break   #2022-03-13

                if mst_level==10:
                    continue
                # Connection
                if mst_flag and slv_flag:
                    #self.con_pin_to_pin(mst_inst, mst_port, mst_msb, mst_lsb, slv_inst, slv_port, slv_msb, slv_lsb)
                    if mst_msb>slv_msb: com_msb = slv_msb
                    else              : com_msb = mst_msb
                    if mst_lsb>slv_lsb: com_lsb = slv_lsb
                    else              : com_lsb = mst_lsb
                    self.con_pin_to_pin(mst_inst, mst_port, com_msb, com_lsb, slv_inst, slv_port, com_msb, com_lsb)

                    # Tie open bits due to bit difference
                    if (mst_dir=='input'):
                        if mst_msb>slv_msb: self.con_pin_value(mst_inst, mst_port, mst_msb, (slv_msb+1), mst_default)#; print ('DEBUG 1 ===>')
                        if mst_lsb<slv_lsb: self.con_pin_value(mst_inst, mst_port, (slv_lsb-1), mst_lsb, mst_default)#; print ('DEBUG 2 ===>')
                    if (slv_dir=='input'):
                        if slv_msb>mst_msb: self.con_pin_value(slv_inst, slv_port, slv_msb, (mst_msb+1), mst_default)#; print ('DEBUG 3 ===>')
                        if slv_lsb<mst_lsb: self.con_pin_value(slv_inst, slv_port, (mst_lsb-1), slv_lsb, mst_default)#; print ('DEBUG 4 ===>')

                elif mst_flag and (mst_dir=='input'):
                    self.con_pin_value(mst_inst, mst_port, mst_msb, mst_lsb, mst_default)
                elif slv_flag and (slv_dir=='input'):
                    if mst_level!=11:
                        self.con_pin_value(slv_inst, slv_port, slv_msb, slv_lsb, mst_default)
                
                # Lint Check
                if mst_flag and slv_flag:
                    if mst_msb!=slv_msb:
                        message = 'The msb values of master and slave are different'
                        if   mst_level<3: self.__rpt_group(con, mst_port, slv_port, 'E', message)
                        else            : self.__rpt_group(con, mst_port, slv_port, 'W', message)
                    if mst_lsb!=slv_lsb:
                        message = 'The lsb values of master and slave are different'
                        if   mst_level<3: self.__rpt_group(con, mst_port, slv_port, 'E', message)
                        else            : self.__rpt_group(con, mst_port, slv_port, 'W', message)
                elif mst_flag:
                    if   mst_dir=='input': message = 'The Master Port is tied because the Slave Port cannot be found'
                    else                 : message = 'Slave Port cannot be found'
                    if   mst_level==0: self.__rpt_group(con, mst_port, slv_port, 'E', message)
                    elif mst_level<2 : self.__rpt_group(con, mst_port, slv_port, 'W', message)
                    else             : self.__rpt_group(con, mst_port, slv_port, 'I', message)
                elif slv_flag:
                    if   slv_dir=='input': message = 'The Slave Port is tied because the Master Port cannot be found'
                    else                 : message = 'Master Port cannot be found'
                    if   mst_level==0: self.__rpt_group(con, mst_port, slv_port, 'E', message)
                    elif mst_level<2 : self.__rpt_group(con, mst_port, slv_port, 'W', message)
                    else             : self.__rpt_group(con, mst_port, slv_port, 'I', message)
                    #-== if mst_level==11 and (mst_dir=='input'):
                    #-==     if hready_loopback!='':
                    #-==         self.con_pin_to_pin(mst_inst, hready_loopback, 0, 0, slv_inst, slv_port, slv_msb, slv_lsb)
                    #-==     else:
                    #-==         print('[LOG] *E, Cannot connect HREADY Signal <================== ERROR')
                else:
                    message = 'The Master/Slave Port cannot be found'
                    if   mst_level==0: self.__rpt_group(con, mst_port, slv_port, 'E', message)
                    elif mst_level==1: self.__rpt_group(con, mst_port, slv_port, 'W', message)
                    elif mst_level<3 :
                        #print ('--------')
                        #print (con)
                        self.__rpt_group(con, mst_port, slv_port, 'I', message)

            self.connect_hready(dic_hready)


    def connect_hready(self, dic):
        mst_inst = dic['mst']['inst']
        slv_inst = dic['slv']['inst']
        #if mst_inst=='' and slv_inst=='':
        #    return True
        #print (dic)
        try:
            self.DTOP[mst_inst][dic['mst']['input']]
            mst_i_port = dic['mst']['input']
            mst_i_flag = True
        except:
            mst_i_flag = False
        try:
            self.DTOP[mst_inst][dic['mst']['output']]
            mst_o_port = dic['mst']['output']
            mst_o_flag = True
        except:
            mst_o_flag = False
        try:
            self.DTOP[slv_inst][dic['slv']['input']]
            slv_i_port = dic['slv']['input']
            slv_i_flag = True
        except:
            slv_i_flag = False
        try:
            self.DTOP[slv_inst][dic['slv']['output']]
            slv_o_port = dic['slv']['output']
            slv_o_flag = True
        except:
            slv_o_flag = False

        if mst_i_flag==True and mst_o_flag==False and slv_i_flag==False and slv_o_flag==True:
            self.con_pin_to_pin(mst_inst, mst_i_port, 0, 0, slv_inst, slv_o_port, 0, 0)
        elif mst_i_flag==True and mst_o_flag==False and slv_i_flag==True and slv_o_flag==True:
            self.con_pin_to_pin(mst_inst, mst_i_port, 0, 0, slv_inst, slv_o_port, 0, 0)
            self.con_pin_to_pin(mst_inst, mst_i_port, 0, 0, slv_inst, slv_i_port, 0, 0)
        elif mst_i_flag==True and mst_o_flag==True and slv_i_flag==False and slv_o_flag==True:
            self.con_pin_to_pin(mst_inst, mst_i_port, 0, 0, slv_inst, slv_o_port, 0, 0)
            print ('[LOG] *E, Cannot connect master hready output port ')
            print ('          - master instance name = ', mst_inst)
            print ('               ==> hready output = ', mst_o_port)
            print ('               ==> hready input  = ', mst_i_port)
            print ('          - slave instance name  = ', slv_inst)
            print ('               ==> hready output = ', slv_o_port)
        elif mst_i_flag==True and mst_o_flag==True and slv_i_flag==True and slv_o_flag==True:
            self.con_pin_to_pin(mst_inst, mst_i_port, 0, 0, slv_inst, slv_o_port, 0, 0)
            self.con_pin_to_pin(mst_inst, mst_o_port, 0, 0, slv_inst, slv_i_port, 0, 0)
        else:
            return True



    def __rpt_group(self, icon, imst_port, islv_port, ilevel, imessage):
        #self.DEBUG = 0:off *E, 1: off *W, 2: off INFO, 3: all
        if   ilevel=='E':
            if self.DEBUG>0: lvl = '*E'
            else           : return True
        elif ilevel=='W':
            if self.DEBUG>1: lvl = '*W'
            else           : return True
        elif ilevel=='I':
            if self.DEBUG>2: lvl = 'INFO'
            else           : return True
        else:
            assert(0), '[ERROR] Unknown level'
        
        print ('[LOG] %s, bus_group : %s'%(lvl, imessage))
        try:
            mst_inst = icon['mst']['inst']
            mst_msb = self.DTOP[mst_inst][imst_port]['msb']
            mst_lsb = self.DTOP[mst_inst][imst_port]['lsb']
            print ('     - master inst_name : %s, port_name : %s, msb : %d, lsb : %d'%(mst_inst, imst_port, mst_msb, mst_lsb))
        except:
            print ('     - master inst_name : %s, port_name : %s'%(mst_inst, imst_port))
        try:
            slv_inst = icon['slv']['inst']
            slv_msb = self.DTOP[slv_inst][islv_port]['msb']
            slv_lsb = self.DTOP[slv_inst][islv_port]['lsb']
            print ('     - slave  inst_name : %s, port_name : %s, msb : %d, lsb : %d'%(slv_inst, islv_port, slv_msb, slv_lsb))
        except:
            print ('     - slave  inst_name : %s, port_name : %s'%(slv_inst, islv_port))
        return True

#    def is_digit (self, istr):
#        istr = istr.split('\'')
#        if istr[0].isdigit() or istr[0].strip()==''
#            return True
#        else:
#            return False


    def connection_pattern (self):
        for con in self.JSON['connection']:
            if con['type']!='pattern':continue
            # Pattern Operation
            mst_inst = con['mst']['inst']
            mst_sym  = con['mst']['symbol']
            slv_inst = con['slv']['inst']
            slv_sym  = con['slv']['symbol']
            mst_mod = self.JSON['instance'][mst_inst]['modname']

            # no 'tie' mode
            if mst_inst==''    : mode = 't2t'
            #    if slv_inst=='': mode = 't2t'
            #    else           : x
            else:
                if slv_inst=='': mode = 'p2t'
                else           : mode = 'p2p'
            
            list_con = []
            if mode=='p2p' or 'p2t':
                for dic_port in self.JSON['module'][mst_mod]['ports']:
                    # find master port
                    flag_hit, mst_port_common = self.pattern_mst_hit(dic_port, mst_sym)
                    if flag_hit==False:
                        continue
                    # find slave port
                    if mode=='p2t': slave_inst_name = '$top'
                    else          : slave_inst_name = slv_inst
                    slv_port = self.pattern_slv_port (mst_port_common, dic_port['dir'], slave_inst_name, slv_sym)
                    if self.DTOP[slave_inst_name].get(slv_port,'not_found')=='not_found':
                        continue
                    #dic_con = self.connection_pattern_gen(mode, dic_port, slv_inst, slv_port)
                    list_con.append(self.gen_from_connection_pattern(mode, mst_inst, dic_port, slv_inst, slv_port))
            # t2t mode
            else:
                for dic_port in self.JSON['top']['ports']:
                    # find master port
                    flag_hit, mst_port_common = self.pattern_mst_hit(dic_port, mst_sym)
                    if flag_hit==False:
                        continue
                    # find slave port
                    slv_port = self.pattern_slv_port (mst_port_common, dic_port['dir'], '$top', slv_sym)
                    if self.DTOP[slv_inst].get(slv_port,'not_found')=='not_found':
                        continue
                    list_con.append(self.gen_from_connection_pattern(mode, mst_inst, dic_port, slv_inst, slv_port))
            # Append Connection
            for con in list_con:
                if len(con)==0: continue
                self.JSON['connection'].append(con)

    def gen_from_connection_pattern(self, mode, mst_inst, dic_port, slv_inst, slv_port):
#        if mode=='t2t': mst_inst_name = '$top'
#        else          : mst_inst_name = mst_inst
        if mode=='p2t': slv_inst_name = '$top'
        else          : slv_inst_name = slv_inst
        
        mst_port = dic_port['port']
        mst_msb  = dic_port['msb'];  slv_msb = self.DTOP[slv_inst_name][slv_port]['msb']
        mst_lsb  = dic_port['lsb'];  slv_lsb = self.DTOP[slv_inst_name][slv_port]['lsb']
        mst_dir  = dic_port['dir'];  slv_dir = self.DTOP[slv_inst_name][slv_port]['dir']        
        result = {}
        result['type'] = mode
        if mode=='p2p':
            if (mst_dir=='inout' and mst_dir==slv_dir) or (mst_dir!='inout' and mst_dir!=slv_dir):
                result['type'] = 'p2p'
                result['mst']={'inst':mst_inst,'port':mst_port,'msb':mst_msb,'lsb':mst_lsb}
                result['slv']={'inst':slv_inst,'port':slv_port,'msb':slv_msb,'lsb':slv_lsb}
        elif mode=='p2t':
            if mst_dir==slv_dir:    # only same direction
                result['type'] = 'p2t'
                result['mst']={'inst':mst_inst,'port':mst_port,'msb':mst_msb,'lsb':mst_lsb}
                result['top']={                'port':slv_port,'msb':slv_msb,'lsb':slv_lsb}
        elif mode=='t2t':
            if (mst_dir=='inout' and mst_dir==slv_dir) or (mst_dir!='inout' and mst_dir!=slv_dir):
                result['type'] = 't2t'
                result['mst']={'inst':'','port':mst_port,'msb':mst_msb,'lsb':mst_lsb}
                result['slv']={'inst':'','port':slv_port,'msb':slv_msb,'lsb':slv_lsb}
        return result


#            list_con = []
#            for dic_port in self.JSON['module'][mst_mod]['ports']:
#                flag_hit, mst_port_common = self.pattern_mst_hit(dic_port, mst_sym)
#                if flag_hit:
#                    if slv_inst=='':
#                        slv_inst = '$top'
#                    
#                    slv_port = self.pattern_slv_port (mst_port_common, dic_port['dir'], slv_inst, slv_sym)
#                    if self.DTOP[slv_inst].get(slv_port,'not_found')=='not_found':
#                        continue
#
#                    mst_port = dic_port['port']
#                    mst_msb  = dic_port['msb'];  slv_msb = self.DTOP[slv_inst][slv_port]['msb']
#                    mst_lsb  = dic_port['lsb'];  slv_lsb = self.DTOP[slv_inst][slv_port]['lsb']
#                    mst_dir  = dic_port['dir'];  slv_dir = self.DTOP[slv_inst][slv_port]['dir']
#                    result = {}
#                    #if slv_inst=='' or slv_inst=='$top':
#                    if mode=='p2t':
#                        result['type'] = 'p2t'
#                        result['mst']={'inst':mst_inst,'port':mst_port,'msb':mst_msb,'lsb':mst_lsb}
#                        result['top']={                'port':slv_port,'msb':slv_msb,'lsb':slv_lsb}
#                        if mst_dir==slv_dir:    # only same direction
#                            list_con.append(result)
#                    elif mode=='p2p':
#                        result['type'] = 'p2p'
#                        result['mst']={'inst':mst_inst,'port':mst_port,'msb':mst_msb,'lsb':mst_lsb}
#                        result['slv']={'inst':slv_inst,'port':slv_port,'msb':slv_msb,'lsb':slv_lsb}
#                        if (mst_dir=='inout' and mst_dir==slv_dir) or (mst_dir!='inout' and mst_dir!=slv_dir):
#                            list_con.append(result)
#                    else:
#                        print ('t2t')
#            # Append Connection
#            for con in list_con:
#                self.JSON['connection'].append(con)

    # for pattern type - prefix dir
    def __pdir(self, idir):
        if   idir=='input' : return 'i_'
        elif idir=='output': return 'o_'
        elif idir=='inout' : return 'b_'
        else: assert(0), '[ERROR] Unknown direction = %s'%(idir)
    # for pattern type - surfix dir
    def __sdir(self, idir):
        if   idir=='input' : return '_i'
        elif idir=='output': return '_o'
        elif idir=='inout' : return '_b'
        else: assert(0), '[ERROR] Unknown direction = %s'%(idir)
    # for pattern type - invert dir
    def __idir(self, idir):
        if   idir=='i_': return 'o_'
        elif idir=='o_': return 'i_'
        elif idir=='_i': return '_o'
        elif idir=='_o': return '_i'
        else           : return idir

    # for pattern
    def pattern_mst_hit(self, dic_port, symbol):
        port    = dic_port['port']
        if symbol=='$$':
            return True, port
        else:
            symbol = symbol.split('$')
            dir     = dic_port['dir']
        # prefix process
        # --> remove pre_dir
        flag_pre = False
        if symbol[1].startswith('.'):
            prefix = symbol[1][1:]
            if port.startswith(self.__pdir(dir)):
                flag_pre = True
                port = port[2:]
        else:
            flag_pre = True
            prefix = symbol[1]
        # --> remove prefix
        if flag_pre:
            len_width = len(prefix)
            if len_width==0:
                flag_pre = True
            elif port.startswith(prefix):
                flag_pre = True
                port = port[len_width:]
            else:
                flag_pre = False
        if flag_pre==False: return False, ''

        # postfix process
        flag_post = False
        if symbol[2].startswith('.'):
            postfix = symbol[2][1:]
            if port.endswith(self.__sdir(dir)):
                flag_post = True
                port = port[:-2]
        else:
            flag_post = True
            postfix = symbol[2]

        if flag_post:
            len_width = len(postfix)
            if len_width==0:
                flag_post = True
            elif port.endswith(postfix):
                flag_post = True
                port = port[:-len_width]
            else:
                flag_post = False

        if flag_post: return True, port
        else        : return False, ''

    def pattern_slv_port(self, mst_port, mst_dir, slv_inst, slv_symbol):
        slv_symbol = slv_symbol.split('$')
        port = mst_port
        if slv_inst=='' or slv_inst=='$top': is_top = True
        else                               : is_top = False
        # Prefix - process
        if slv_symbol[1].startswith('.'):   # add prefix_port
#            port  = slv_symbol[1][1:]+port  # add prefix
#            affix = self.__pdir(mst_dir)    # add pre_dir
#            if slv_inst!='':                # toggle dir
#                affix = self.__tdir(affix)  # convert port dir
#            port = affix+port

            port  = slv_symbol[1][1:]+port  # add prefix
            if is_top : affix = self.__pdir(mst_dir)
            else      : affix = self.__pdir(self.__idir(mst_dir))
            port = affix+port
        else:
            port = slv_symbol[1]+port

        # Postfix - process
        if slv_symbol[2].startswith('.'):   # add sufix_port
#            port  = port+slv_symbol[2][1:]  # add prefix
#            affix = self.__sdir(mst_dir)    # add pre_dir
#            if slv_inst!='':                # toggle dir
#                affix = self.__tdir(affix)  # convert port dir
#            port = port+affix

            port  = port+slv_symbol[2][1:]  # add prefix
            if is_top : affix = self.__sdir(mst_dir)    # add pre_dir
            else      : affix = self.__sdir(self.__idir(mst_dir))    # add pre_dir
            port = port+affix
        else:
            port = port+slv_symbol[2]

        return port


    def dtop_set_role_con(self):
        for con in self.JSON['connection']:
            if con['type']=='bus_group' or con['type']=='pattern' or con['type']=='alp_vip':
                continue
            mst_inst = con['mst']['inst']
            mst_port = con['mst']['port']
            mst_dir  = self.DTOP[mst_inst][mst_port]['dir']
            if (con['type']=='p2p'):
                slv_inst = con['slv']['inst']
                slv_port = con['slv']['port']
                slv_dir  = self.DTOP[slv_inst][slv_port]['dir']
            elif (con['type']=='p2t'):
                slv_port = con['top']['port']
            elif (con['type']=='tie'):
                tie_value = con['value']
            #--------------------------------------------------------
            # Role
            #--------------------------------------------------------
            if (con['type']=='p2p'):
                if (self.DTOP[mst_inst][mst_port]['role']==''):
                    if (mst_inst!=''): self.DTOP[mst_inst][mst_port]['role'] = 'mst'
                else:
#todo  mst - slv ==> mst
                    #-== if (self.DTOP[mst_inst][mst_port]['role'] == 'slv'):
                    #-==     if (mst_inst!=''): self.DTOP[mst_inst][mst_port]['role'] = 'mst'

                    if (self.DTOP[mst_inst][mst_port]['role'] == 'slv'):
                        print('[LOG] *E, 001, mst_inst = %s, mst_port = %s'%(mst_inst, mst_port))
                        #print(con)


                if (self.DTOP[slv_inst][slv_port]['role']==''):
                    if (slv_inst!=''): self.DTOP[slv_inst][slv_port]['role'] = 'slv'
                else:
#todo slv - mst ==> mst
                    #if (self.DTOP[slv_inst][slv_port]['role'] == 'mst'):
                    #    if (slv_inst!=''): self.DTOP[slv_inst][slv_port]['role'] = 'mst'
                    if (self.DTOP[slv_inst][slv_port]['role'] == 'mst'):
                        print('[LOG] *E, 002, mst_inst = %s, mst_port = %s'%(mst_inst, mst_port))
                        #print(con)
            #--------------------------------------------------------
            # Connection
            #--------------------------------------------------------
            port_msb = self.DTOP[mst_inst][mst_port]['msb']
            port_lsb = self.DTOP[mst_inst][mst_port]['lsb']
            con_msb  = con['mst']['msb'] 
            con_lsb  = con['mst']['lsb'] 
            mst_lsb = con['mst']['lsb']
            #--------------------------------------------------------
            # register mst connection ==> p2p/p2t/tie case
            if (self.DTOP[mst_inst][mst_port]['con'].get(mst_lsb,'new_bit')=='new_bit'):
                self.DTOP[mst_inst][mst_port]['con'][mst_lsb] = []
            # multi input connection (input and con[mst_lsb]>0) ==> bypass
            if con['type']=='p2t' and mst_dir=='input' and len(self.DTOP[mst_inst][mst_port]['con'][mst_lsb])>0:
                if con['top']['port']!=self.DTOP[mst_inst][mst_port]['con'][mst_lsb][0]['port']:
                    assert(0), "[ERROR] Multi input connection, alpgen.dtop_set_role_con\n %s \n %s"%(con, self.DTOP[mst_inst][mst_port]['con'][mst_lsb])
            else:
                self.DTOP[mst_inst][mst_port]['con'][mst_lsb].append(self.__get_dtop_con('mst', con))
            #--------------------------------------------------------
            # register slv connection ==> p2p/p2t case
            if (con['type']=='p2p'):
                slv_lsb = con['slv']['lsb']
                if (self.DTOP[slv_inst][slv_port]['con'].get(slv_lsb,'new_bit')=='new_bit'):
                    self.DTOP[slv_inst][slv_port]['con'][slv_lsb] = []
                if slv_dir=='input' and len(self.DTOP[slv_inst][slv_port]['con'][slv_lsb])>0:
                    pass # multi-input
                else:
                    self.DTOP[slv_inst][slv_port]['con'][slv_lsb].append(self.__get_dtop_con('slv', con))

            elif (con['type']=='p2t'):
                slv_lsb = con['top']['lsb']
                if (self.DTOP['$top'][slv_port]['con'].get(slv_lsb,'new_bit')=='new_bit'):
                    self.DTOP['$top'][slv_port]['con'][slv_lsb] = []
                self.DTOP['$top'][slv_port]['con'][slv_lsb].append(self.__get_dtop_con('slv', con))

    def dtop_set_wire(self):
        #ENB_DEBUG_CON = False
        for inst_name in self.JSON['instance'].keys():
            mod_name = self.JSON['instance'][inst_name]['modname']
            #for port in self.JSON['instance'][inst_name]['ports']:
            for port in self.JSON['module'][mod_name]['ports']:
                port_name = port['port']
                port_dir  = self.DTOP[inst_name][port_name]['dir']
                port_msb  = self.DTOP[inst_name][port_name]['msb']
                port_lsb  = self.DTOP[inst_name][port_name]['lsb']
                port_role = self.DTOP[inst_name][port_name]['role']
                port_con  = self.DTOP[inst_name][port_name]['con'] #{0: [{'width': 1, 'port': 'i_clock', 'msb': 0, 'lsb': 0, 'dir': 'input'}]}
                # Top Port Only
                if port_role=='':
                    wire_name = self.__gen_wire_name_from_top(inst_name, port_name)
                # Master (except Top only)
                elif port_role=='mst':
                    wire_name = self.__gen_wire_name_from_mst(inst_name, port_name)
                # Slave (except Top_only)
                elif port_role=='slv':
                    wire_name = self.__gen_wire_name_from_slv(inst_name, port_name)
                else:
                    print('[LOG] *E, 100')
                # write wire name
                self.DTOP[inst_name][port_name]['wire'] = wire_name

    def __gen_wire_name_from_top(self, inst_name, port_name):
        port_dir  = self.DTOP[inst_name][port_name]['dir']
        port_msb  = self.DTOP[inst_name][port_name]['msb']
        port_lsb  = self.DTOP[inst_name][port_name]['lsb']
        port_role = self.DTOP[inst_name][port_name]['role']
        port_con  = self.DTOP[inst_name][port_name]['con']

        if   port_dir=='output': open_name = self.WIRE_NAME_OPEN_O
        elif port_dir=='input' : open_name = self.WIRE_NAME_OPEN_I
        elif port_dir=='inout' : open_name = self.WIRE_NAME_OPEN_B
        else:
            assert(0), '[ERROR], Unknown Direction'

        l_wire_name = []
        pre_lsb = port_lsb
        for bit_num in range(port_lsb, port_msb+1, 1):
            if (port_con.get(bit_num, 'not_found')=='not_found'):
                continue
            for con in port_con[bit_num]:
                open_width = bit_num - pre_lsb
                if (open_width>0):
                    open_lsb = pre_lsb
                    open_msb = bit_num - 1
                    open_lsb = (open_lsb - pre_lsb)
                    open_msb = (open_msb - pre_lsb)

                    if (port_dir=='output'):
                        open_lsb = self.open_o + open_lsb
                        open_msb = self.open_o + open_msb
                        self.open_o = self.open_o + open_width
                    elif (port_dir=='input'):
                        open_lsb = self.open_i + open_lsb
                        open_msb = self.open_i + open_msb
                        self.open_i = self.open_i + open_width
                        print ('[LOG] *W, The Input Port is open')
                        print ('       - instance name = ', inst_name)
                        print ('       - port name     = ', port_name)
                        print ('       - lsb bit       = ', bit_num)
                    elif (port_dir=='inout'):
                        open_lsb = self.open_b + open_lsb
                        open_msb = self.open_b + open_msb
                        self.open_b = self.open_b + open_width

                    if (open_msb==open_lsb): l_wire_name.insert(0, '{0}[{1}]'.format(open_name, open_msb))
                    else                   : l_wire_name.insert(0, '{0}[{1}:{2}]'.format(open_name, open_msb, open_lsb))
                pre_lsb = bit_num + con['width']
                # top port name
                if con['type']=='p2t':
                    top_port = con['port']
                    top_width = self.DTOP['$top'][top_port]['msb']-self.DTOP['$top'][top_port]['lsb']+1
                    con_width = con['width']

                    top_lsb  = self.DTOP['$top'][top_port]['lsb']
                    top_msb  = self.DTOP['$top'][top_port]['msb']

                    con_msb  = int(con['lsb']) + con['width'] - 1
                    #con['lsb']==top_lsb) and (con_msb==top_msb):

                    if (top_width==con_width):
                        wire_name = top_port
                    else:
                        con_msb = con['lsb']+con['width']-1
                        if(con_msb==con['lsb']): wire_name = top_port+'['+str(con_msb)+']' #top 
                        else                   : wire_name = top_port+'['+str(con_msb)+':'+str(con['lsb'])+']' #top 
                    l_wire_name.insert(0, wire_name)
                elif con['type']=='tie':
                    wire_name = self.conv_dec_to_hex(con['value'], con['width'])
                    l_wire_name.insert(0, wire_name)
                else:
                    print ('[LOG] *E, There can be no conditions other than p2t, tie')
                    print ('           - type          = ', con['type'])
                    print ('           - instance name = ', inst_name)
                    print ('           - port name     = ', port_name)
                    print ('           - con           = ', con)
                    exit()
        # check msb open bit 
        open_width = (port_msb+1) - pre_lsb
        if (open_width>0):
            #l_wire_name.insert(0, '{0}\'b0'.format(num_open))
            open_lsb = pre_lsb
            open_msb = port_msb
            open_lsb = (open_lsb - pre_lsb)
            open_msb = (open_msb - pre_lsb)
            if (port_dir=='output'):
                open_lsb = self.open_o + open_lsb
                open_msb = self.open_o + open_msb
                self.open_o = self.open_o + open_width
            elif (port_dir=='input'):
                open_lsb = self.open_i + open_lsb
                open_msb = self.open_i + open_msb
                self.open_i = self.open_i + open_width
            elif (port_dir=='inout'):
                open_lsb = self.open_b + open_lsb
                open_msb = self.open_b + open_msb
                self.open_b = self.open_b + open_width

            if (open_msb==open_lsb): l_wire_name.insert(0, '{0}[{1}]'.format(open_name, open_msb))
            else                   : l_wire_name.insert(0, '{0}[{1}:{2}]'.format(open_name, open_msb, open_lsb))

        _len = len(l_wire_name)
        if   (_len==0): return ''
        elif (_len==1): return  ','.join(l_wire_name)
        else          : return  '{'+','.join(l_wire_name)+'}'


    def __gen_wire_name_from_mst(self, inst_i, port_i):
        port_name = self.__get_wire_name(inst_i, port_i, '', '')
        port_dir  = self.DTOP[inst_i][port_i]['dir']
        port_msb  = self.DTOP[inst_i][port_i]['msb']
        port_lsb  = self.DTOP[inst_i][port_i]['lsb']
        port_con  = self.DTOP[inst_i][port_i]['con']
        port_width = port_msb - port_lsb + 1

        self.__reg_wire_mst(port_name, port_msb, port_lsb)

        for bit_num in range(port_lsb, port_msb+1, 1):
            if (port_con.get(bit_num, 'not_found')=='not_found'): continue
            # only p2t/tie is processed
            for con in port_con[bit_num]:
                bit_width = con['width']
                if (con['type']=='p2t'):
                    con_msb = con['lsb']+con['width']-1
                    bit_msb = bit_num + con['width']-1
                    if (port_dir=='input'):
                        #2020-02-27
                        #self.__reg_assign(port_name, bit_msb, bit_num, con['port'], con_msb, con['lsb'])
                        self.__reg_assign(port_name, port_width, bit_msb, bit_num, con['port'], con['width'], con_msb, con['lsb'])
                    elif (port_dir=='output'):
                        self.__reg_assign(con['port'], con['width'], con_msb, con['lsb'], port_name, port_width, bit_msb, bit_num)
                    else:
                        print('[LOG] *E, not support')
                elif (con['type']=='tie'):
                    if (port_dir=='input'):
                        bit_msb = bit_num + con['width']-1
                        # 2022-02-27
                        #self.__reg_assign_tie(port_name, bit_msb, bit_num, con['value'])
                        self.__reg_assign_tie(port_name, port_width, bit_msb, bit_num, con['value'])
                        # if port_name=="cpu_i_hresp":
                        #     print('---------')
                        #     print (port_name)
                        #     print (port_width)
                        #     print (bit_width)
                        #     print (bit_msb)
                        #     print (bit_num)
                        #     print (con['value'])
                        #     exit()

                    else:
                        print('[LOG] *E, not support')
        return port_name

    def __gen_wire_name_from_slv(self, inst_name, port_name):
        port_dir  = self.DTOP[inst_name][port_name]['dir']
        port_msb  = self.DTOP[inst_name][port_name]['msb']
        port_lsb  = self.DTOP[inst_name][port_name]['lsb']
        port_con  = self.DTOP[inst_name][port_name]['con']
        slv_width = port_msb - port_lsb + 1

        if (port_dir=='inout'  and len(port_con)==1) or \
           (port_dir=='input'  and len(port_con)==1) or \
           (port_dir=='output' and len(port_con)==1 and len(list(port_con.values())[0])==1):
            bit_num = list(port_con.keys())[0]
            con = port_con[bit_num][0]
            mst_inst  = con['inst']
            mst_port  = con['port']
            mst_lsb   = con['lsb']
            mst_msb   = con['lsb']+con['width']-1
            #con_width = con['width']
            # write wire name on the basis of master
            if mst_msb==self.DTOP[mst_inst][mst_port]['msb'] and mst_lsb==self.DTOP[mst_inst][mst_port]['lsb']:
                return self.__get_wire_name(mst_inst, mst_port, '', '')
            else:
                return self.__get_wire_name(mst_inst, mst_port, mst_msb, mst_lsb)
        else:
            if (port_dir=='inout'):
                print ('[LOG] *E, The inout port does not support multi-connection.')
                exit('==> Error')

            wire_name_slv = self.__get_wire_name(inst_name, port_name, '', '')     # slave
            self.__reg_wire_slv(wire_name_slv, port_msb, port_lsb)

            for bit_num in range(port_lsb, port_msb+1, 1):
                if (port_con.get(bit_num, 'not_found')=='not_found'):
                    continue
                for con in port_con[bit_num]:
                    slv_lsb = bit_num
                    slv_msb = bit_num+con['width']-1
                    if (con['type']=='p2p'):
                        mst_lsb = con['lsb']
                        mst_msb = con['lsb']+con['width']-1
                        mst_width = con['width']
                        wire_name_mst = self.__get_wire_name(con['inst'], con['port'], '', '') # master
                        if (port_dir=='input'):
                            #2020-02-27
                            #self.__reg_assign(wire_name_slv, slv_msb, slv_lsb, wire_name_mst, mst_msb, mst_lsb)
                            self.__reg_assign(wire_name_slv, slv_width, slv_msb, slv_lsb, wire_name_mst, mst_width, mst_msb, mst_lsb)
                        elif (port_dir=='output'):
                            self.__reg_assign(wire_name_mst, mst_width, mst_msb, mst_lsb, wire_name_slv, slv_width, slv_msb, slv_lsb)
                    elif (con['type']=='p2t'):
                        # slv_lsb = bit_num
                        # slv_msb = bit_num+con['width']-1
                        # mst_lsb = con['lsb']
                        # mst_msb = con['lsb']+con['width']-1
                        mst_lsb = con['lsb']
                        mst_msb = con['lsb']+con['width']-1
                        mst_width = con['width']
                        wire_name_mst = con['port']  # top port
                        if (port_dir=='input'):
                            #2020-02-27
                            self.__reg_assign(wire_name_slv, slv_width, slv_msb, slv_lsb, wire_name_mst, mst_width, mst_msb, mst_lsb)
                        elif (port_dir=='output'):
                            self.__reg_assign(wire_name_mst, mst_width, mst_msb, mst_lsb, wire_name_slv, slv_width, slv_msb, slv_lsb)
                    elif (con['type']=='tie'):
                        # slv_lsb = bit_num
                        # slv_msb = bit_num+con['width']-1
                        if (port_dir=='input'):
                            self.__reg_assign_tie(wire_name_slv, slv_width, slv_msb, slv_lsb, con['value'])
                            if wire_name_slv=="cpu_i_hresp":
                                print('====================')
                                print (wire_name_slv)
                                print (slv_width)
                                print (slv_msb)
                                print (slv_lsb)
                                print (con['value'])
                                exit()


                        else:
                            print('[LOG] no condition')
                            exit()

            # write wire name on the basis of slave. becase of con>1
            return wire_name_slv 
    
    # con = {"inst": instance_name, "port": port_name, "msb": msb_bit, "lsb": lsb_bit, "dir": output}
    def __get_dtop_con (self, itype, con):
        dic = {}
        dic['width'] = con['mst']['msb']-con['mst']['lsb']+1   # mst_width==slv_width
        # Slave interface (return Master Info)
        if (itype=='slv'):
            if  not (con['type']=='p2p' or con['type']=='p2t' or con=='tie'):
                assert(0), '[LOG] *E, Unknown con[\'type\'] = '+con['type']
            dic['type'] = con['type']
            dic['inst'] = con['mst']['inst']
            dic['port'] = con['mst']['port']
            dic['lsb']  = con['mst']['lsb']
            dic['dir']  = self.DTOP[dic['inst']][dic['port']]['dir']
        # Master interface (return Slave Info)
        elif (itype=='mst'):
            # Corresponding Port Width
            dic['type'] = con['type']
            # Connected Port Info
            if(con['type']=='p2p'):
                dic['inst'] = con['slv']['inst']
                dic['port'] = con['slv']['port']
                dic['lsb']  = con['slv']['lsb']
                dic['dir']  = self.DTOP[dic['inst']][dic['port']]['dir']
            elif (con['type']=='p2t'):
                dic['port']  = con['top']['port']
                dic['lsb']   = con['top']['lsb']
                dic['dir']  = self.DTOP['$top'][dic['port']]['dir']
            elif (con['type']=='tie'):
                dic['value'] = con['value']
            else:
                assert(0), '[LOG] *E, Unknown con[\'type\'] = '+con['type']
        else:
            assert(0), '[LOG] *E, Unknown itype = '+itype
        return dic

    # convert parameter bit (msb/lsb) - string to int
    def dec_bit(self, ibit, iinst, iport, idebug):
        if type(ibit)==str:
            try:
                return eval(ibit, {}, self.JSON['instance'][iinst]['parameter'])
            except:
                #assert(0), '[ERROR] Cannot find parameter, instance = %s, port = %s, bit = %s, parm = %s'%(iinst, iport, str(ibit), self.JSON['instance'][iinst]['parameter'])
                if self.DEBUG>1 and idebug:
                    print ('[LOG] *W, Cannot find parameter, instance = %s, port = %s, bit = %s'%(iinst, iport, str(ibit)))
                    print ('          - If there is no alp_vip port, there may be no parameter')
                return 0


        else:
            return ibit

    # make port dic
    def dtop_port_format(self, dic_port):
        return {   'role': '', \
                    'bus_group' : False, \
                    'dir' : dic_port['dir'], \
                    'msb' : dic_port['msb'], \
                    'lsb' : dic_port['lsb'], \
                    'con_inst' : '', \
                    'con_port' : '', \
                    'wire': '', \
                    'con': {} }

    #def dtop_port_format1(self, dic_port, dic_param, iinst):
    def dtop_port_format1(self, dic_port, iinst):
        tmp = {}
        tmp['role']      = ''
        tmp['bus_group'] = False
        tmp['wire']      = ''
        tmp['con']       = {}
        tmp['dir']       = dic_port['dir']
        tmp['msb']       = self.dec_bit(dic_port['msb'], iinst, dic_port['port'], False)
        tmp['lsb']       = self.dec_bit(dic_port['lsb'], iinst, dic_port['port'], False)
        return tmp
        #
        #if type(dic_port['msb'])==str:
        #    try:
        #        tmp['msb'] = eval(dic_port['msb'], dic_param)
        #    except:
        #        assert(0), '[ERROR] Cannot find parameter, port_name = %s, msb = %s'%(dic_port['port'], dic_port['msb'])
        #else:
        #    tmp['msb'] = dic_port['msb']
        #if type(dic_port['lsb'])==str:
        #    try:
        #        tmp['lsb'] = eval(dic_port['lsb'], dic_param)
        #    except:
        #        assert(0), '[ERROR] Cannot find parameter, port_name = %s, lsb = %s'%(dic_port['port'], dic_port['lsb'])
        #else:
        #    tmp['lsb'] = dic_port['lsb']
        #return tmp
        

    # make wire name
    # return : 'prefix'+port_nmae
    #    if no msb/lsb info, the only port name is returned. (ex: prefix+port_name)
    #    if there is msb/lsb info, bit field infor is added. (ex: prefix+port_name[1:0])
    def __get_wire_name(self, inst_name, port_name, msb, lsb):
        wire_name = self.JSON['instance'][inst_name]['prefix']+port_name
        if msb=='' and lsb=='':
            return wire_name
        elif msb>=0 and lsb>=0:
            if msb==lsb: return wire_name+'['+str(msb)+']'
            else       : return wire_name+'['+str(msb)+':'+str(lsb)+']'
        else:
            print('[LOG] *E, Unknown get_wire_name input argument')
            print('      - instance name = ', inst_name)
            print('      - port name     = ', port_name)
            print('      - msb           = ', str(msb))
            print('      - lsb           = ', str(lsb))
            exit()

    def __reg_wire_mst(self, wire_name, msb, lsb):
        self.TOP_WIRE_MST.append ({"name": wire_name, "msb": msb, "lsb":lsb})

    def __reg_wire_slv(self, wire_name, msb, lsb):
        self.TOP_WIRE_SLV.append ({"name": wire_name, "msb": msb, "lsb": lsb})

    def __reg_assign(self, assign_name, assign_width, assign_msb, assign_lsb, var_name, var_width, var_msb, var_lsb):
        self.TOP_ASSIGN.append ([assign_name, assign_width, assign_msb, assign_lsb, var_name, var_width, var_msb, var_lsb])

    def __reg_assign_tie(self, assign_name, assign_width, assign_msb, assign_lsb, tie_value):
        self.TOP_ASSIGN.append ([assign_name, assign_width, assign_msb, assign_lsb, tie_value])

    def __make_json_format (self):
        for mod_name in self.JSON['module']:
            if (self.JSON['module'][mod_name].get('symbol_info','no_key')=='no_key'):
                self.JSON['module'][mod_name]['symbol_info'] = {}

